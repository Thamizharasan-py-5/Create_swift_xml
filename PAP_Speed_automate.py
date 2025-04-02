import requests
import pandas as pd
import openpyxl
from openpyxl.styles import Font
import os
import json
import openpyxl as xl;
import sys
from datetime import datetime, timedelta,date
from dateutil import relativedelta
def APICall(api,stDateStr, endDateStr, stDate, endDate, fileSavePath):
    resTxt=[]
    jsondata=''
    monthDays={}
    monthDays['1']='31'
    monthDays['2']='28'
    monthDays['3']='31'
    monthDays['4']='30'
    monthDays['5']='31'
    monthDays['6']='30'
    monthDays['7']='31'
    monthDays['8']='31'
    monthDays['9']='30'
    monthDays['10']='31'
    monthDays['11']='30'
    monthDays['12']='31'
    try:
        start_date = datetime.strptime(stDate, "%Y-%m-%d")
        end_date = datetime.strptime(endDate, "%Y-%m-%d")
        delta = relativedelta.relativedelta(end_date, start_date)
        count=1
        while count<= 12:
            for x in range(int(delta.months)+1):
                stDate1=stDate
                orgStDate=datetime.strptime(stDate1,'%Y-%m-%d')
                noOfDays=monthDays[str(orgStDate.month)]
                one_month = timedelta(days=int(noOfDays)-1)
                one_month_from_today = datetime.strptime(stDate1,'%Y-%m-%d') + one_month
                one_month_from_today_string = one_month_from_today.strftime('%Y-%m-%d')
                endDate1=one_month_from_today_string
                one_day = timedelta(days=1)
                newEndDate1=datetime.strptime(endDate1, "%Y-%m-%d")
                newEndDate=newEndDate1+one_day
                stDate=newEndDate.strftime('%Y-%m-%d')
                if (x==int(delta.months)):
                    endDate1=endDate
                headers = {
                    'email': 'vijayachandran.c@newgen.co',
                    'token': 'afaa7b95fa0d6ea5a1cbad45181b0bf9-53d61c9f5860b321636130f695c18f9f',
                    'Content-Type': 'application/json',
                }
         
                json_data = {
                    'publisherId': 1,
                    'startDate': stDate1,
                    'endDate': endDate1,
                }
         
                response = requests.post(
                    'https://pubsub.newgen.co/reportsmicro/reports/delay_analysis_pap_b_report',
                    headers=headers,
                    json=json_data,
                )
                
                jsondata=response.text
                print('Start Date :',stDate1)
                print('End Date :',endDate1)
                resTxt=json.loads(jsondata)
                apipath=fileSavePath+str(count)+'.json'
                filepath=fileSavePath+str(count)+'.xlsx'
                print(filepath)
                resTxtStr=json.dumps(resTxt)
                with open(apipath,'w') as fd:
                    fd.write(resTxtStr)
                    fd.close()
                df_json = pd.read_json(apipath)
                df_json = df_json[['JournalShortCode','Doi','TargetPapSpeed','ProjectCreatedDate','ActualEndDate','TaskName','ProjectManager','taskName1','hrsLate1','timelastOnQuery1','taskName2','hrsLate2','timelastOnQuery2','taskName3','hrsLate3','timelastOnQuery3']]
                df_json.to_excel(filepath)
                count +=1
                if endDate1==endDate:
                    break
            if endDate1==endDate:
                break
    except Exception as exx:
        print (exx)
def fun_excelSave(excelFilesavePath1,xl_filecoll):
    try:
        filepath=excelFilesavePath1+'/'+'Monthwise_Excel_Combined_output.xlsx'
        files = os.listdir(xl_filecoll)
        df = pd.DataFrame()
        for file in range(len(files)):
            if files[file].endswith('.xlsx'):
                file_df = pd.read_excel(xl_filecoll + "//" + files[file],sheet_name="Sheet1")
                file_df[''] = str(files[file])
                df = df.append(file_df, ignore_index=True)
                print(df)
        df = df.set_index('Unnamed: 0')
        df.to_excel(filepath,index=True)
    except Exception as exx:
        print (exx)
    return filepath

def CalculatedDiffCreatedDateAndEstimatedDate(ws,calculatedDates,calculatedDates1):
    aheadDelay=[]
    difDays=[]
    try:
        estimateDateWithPapDays=[]
        createdCount=0
        for row in ws['F0':'F0']:
            for cell in row:
                if createdCount==0:
                    createdCount=createdCount+1
                    continue
                createdDate=cell.value
                if(createdDate==''):
                    estimateDateWithPapDays.append('Blank')
                    createdCount=createdCount+1
                    continue
                
                createdDate1=createdDate.split('T')[0]
                dateSplitted=createdDate1.split('-')
                d0=date(int(dateSplitted[0]),int(dateSplitted[1]),int(dateSplitted[2]))
                d1=calculatedDates[createdCount-1]
                d2=calculatedDates1[createdCount-1]
                deltaTime = d1 - d0
                diff=deltaTime.days
                difDays.append(diff)
                if(diff<1):
                    deltaTime = d2 - d0
                    diff=deltaTime.days
                    difDays.pop()
                    difDays.append(diff)
                    if(diff==0):
                        aheadDelay.append('Ontime')
                    elif(diff<1):
                        aheadDelay.append('Delay')
                    else:
                        aheadDelay.append('AHead')
                else:
                    aheadDelay.append('AHead')
                createdCount=createdCount+1
                
    except Exception as exx:
        print (exx)
    return aheadDelay,difDays

def IncreasePAPDaysToCreatedDate(ws,papSpeedDays):
    increasedDate=[]
    increasedDate1=[]
    calculatedDates=[]
    calculatedDates1=[]
    try:
        createdDateWithPapDays=[]
        createdCount=0
        for row in ws['E0':'E0']:
            for cell in row:
                if createdCount==0:
                    createdCount=createdCount+1
                    continue
                createdDate=cell.value
                if(createdDate==''):
                    createdDateWithPapDays.append('Blank')
                    createdCount=createdCount+1
                    continue
                
                createdDate1=createdDate.split('T')[0]
                dateSplitted=createdDate1.split('-')
                countDays=papSpeedDays[createdCount-1]
                dateAdded=date(int(dateSplitted[0]),int(dateSplitted[1]),int(dateSplitted[2]))
                calculatedDate1=date_by_adding_business_days(dateAdded, int(countDays))
                calculatedDate= dateAdded + timedelta(days=int(countDays))#date_by_adding_business_days(dateAdded, int(countDays))
                increaedDateStr=str(calculatedDate.year)+'-'+str(calculatedDate.month)+'-'+str(calculatedDate.day)
                increaedDateStr1=str(calculatedDate1.year)+'-'+str(calculatedDate1.month)+'-'+str(calculatedDate1.day)
                createdCount=createdCount+1
                #createdDateWithPapDays.append()
                increasedDate.append(increaedDateStr)
                increasedDate1.append(increaedDateStr1)
                calculatedDates.append(calculatedDate)
                calculatedDates1.append(calculatedDate1)
    except Exception as ex:
        print (ex)
    return calculatedDates,increasedDate,calculatedDates1,increasedDate1

def PAPSpeedDays(ws):
    papSpeedData=[]
    try:
        speedCount=0
        for row in ws['D0':'D0']:
            for cell in row:
                if speedCount==0:
                    speedCount=speedCount+1
                    continue
                targetspeed=cell.value
                if(targetspeed==''):
                    papSpeedData.append('Blank')
                    speedCount=speedCount+1
                    continue
                mul=int(targetspeed)*7
                speedCount=speedCount+1
                papSpeedData.append(mul)
    except Exception as ex:
        print (ex)
    return papSpeedData
def date_by_adding_business_days(from_date, add_days):
    business_days_to_add = add_days
    current_date = from_date
    while business_days_to_add > 0:
        current_date += timedelta(days=1)
        weekday = current_date.weekday()
        if weekday >= 5: # sunday = 6
            continue
        business_days_to_add -= 1
    return current_date

def UpdateIncreasedDates1(ws,myArr,increasedDate1):
    cntinc=0
    for row in ws['U0':'U0']:
        try:
            if(len(myArr)<cntinc):
                break
            for cell in row:
                if(len(myArr)<cntinc):
                    break
                if(cntinc==0):
                    cntinc=cntinc+1
                    cell.value='Date except weekend'
                    statTxt=Font(bold=True)
                    cell.font=statTxt
                    continue
                else:
                    if(myArr[cntinc-1]=='Delay'):
                        cell.value=increasedDate1[cntinc-1]
                    cntinc=cntinc+1
        except Exception as ex2:
            print (ex2)

def UpdateIncreasedDates(ws,myArr):
    cntinc=0
    for row in ws['R0':'R0']:
        try:
            if(len(myArr)<cntinc):
                break
            for cell in row:
                if(len(myArr)<cntinc):
                    break
                if(cntinc==0):
                    cntinc=cntinc+1
                    cell.value='CalculatedDate'
                    statTxt=Font(bold=True)
                    cell.font=statTxt
                    continue
                else:
                    cell.value=myArr[cntinc-1]
                    cntinc=cntinc+1
        except Exception as ex2:
            print (ex2)
def DiffDays(ws,myArr):
    cntinc=0
    for row in ws['S0':'S0']:
        try:
            if(len(myArr)<cntinc):
                break
            for cell in row:
                if(len(myArr)<cntinc):
                    break
                if(cntinc==0):
                    cntinc=cntinc+1
                    cell.value='Difference'
                    statTxt=Font(bold=True)
                    cell.font=statTxt
                    continue
                else:
                    myVar=-1*int(myArr[cntinc-1])
                    cell.value=myVar
                    cntinc=cntinc+1
        except Exception as ex2:
            print (ex2)

def WriteLastColumn(ws,myArr):
    cntinc=0
    for row in ws['T0':'T0']:
        try:
            if(len(myArr)<cntinc):
                break
            for cell in row:
                if(len(myArr)<cntinc):
                    break
                if(cntinc==0):
                    cntinc=cntinc+1
                    cell.value='Status'
                    statTxt=Font(bold=True)
                    cell.font=statTxt
                    continue
                else:
                    cell.value=myArr[cntinc-1]
                    cntinc=cntinc+1
        except Exception as ex2:
            print (ex2)
def AppendDelayValues(ws2,ws3,mr,mc):
    cont=1
    for i in range (1, mr + 1):
        inc=0
        for j in range (1, mc + 1):
            cellValue=ws2.cell(row=i,column=20).value
            if(cellValue=='Delay'):
            # reading cell value from source excel file
                c = ws2.cell(row = i, column = j)

            # writing the read value to destination excel file
                ws3.cell(row = cont, column = j).value = c.value
                if(inc==0):
                    inc=inc+1
                    cont=cont+1
def AddSecondSheeetValues(wb,ws1,first_sheet,second_sheet):
    first_sheet = wb.worksheets[0]
    second_sheet = wb.create_sheet('Sheet2')
    for row in first_sheet.iter_rows(values_only=True):
            second_sheet.append(row)
    row_num = 1
    bold_font = Font(bold=True)
    for cell in second_sheet[row_num]:
        cell.font = bold_font
    for row in second_sheet['T0':'T0']:
        try:
            for cell in row:
                if cell.value=='AHead' or cell.value=='Ontime':
                    second_sheet.delete_rows(row[0].row,1)
        except Exception as ex:
            print(ex)

def AddSheetinNewExcel(ws,ws1,firstValue,secondValue):
    
    
    myBArr=[]
    count=0
    for row in ws[firstValue:firstValue]:
      for jcode in row:
          try:
              if count==0:
                  count=count+1
                  continue
              myBArr.append(str(jcode.value))        
              count=count+1
          except Exception as ex1:
            print(ex1)
    count1=0
    headingColName=''
    for row1 in ws1[secondValue:secondValue]:
      for jcode1 in row1:
          try:
              if(count1-1>=len(myBArr)):
                 break
              if(count1==0):
                  headingColName=jcode1.value
                  count1=count1+1
                  continue
              colValue=myBArr[count1-1]
              if('T' in colValue and ('Created' in headingColName or 'End' in headingColName or 'Date' in headingColName)):
                  colValue=colValue.split('T')[0]
              jcode1.value=colValue
              count1=count1+1
          except Exception as ex:
              print(ex)

def ChangeNewXlsheetHeading():
    ws1.cell(row=1, column=13).value = 'hrsLate1'
    ws1.cell(row=1, column=16).value = 'hrsLate2'         
    ws1.cell(row=1, column=19).value = 'hrsLate3'
sysPath=sys.argv[0]
exePath=os.path.dirname(sysPath)
stDate=sys.argv[1]
stDate=stDate.replace('_',' ')
endDate=sys.argv[2]
endDate=endDate.replace('_',' ')
excelfile='//j-fs01/OUP_Journals-L/OUP_File_Autodownload/OUPJ-Running-Order-Report/Config_file/PAP_Speed_Sample_Formula.xlsx'
first_sheet=[]
second_sheet=[]
api=('https://pubsub.newgen.co/reportsmicro/reports/delay_analysis_pap_b_report')
filename=api.split('/')
fileTxt=filename[len(filename)-1]
excelSavepath=exePath+'/'+'Excel_Files'+'/'+'Combined_Excel_files'+'/'+fileTxt+'.xlsx'
fileSavePath=exePath+'/'+'Excel_Files'+'/'+fileTxt
excelFilesavePath1=exePath+'/'+'Excel_Files'+'/'+'Combined_Excel_files'
xl_filecoll=exePath+'/'+'Excel_Files'+'/'
delay_xlfile=excelFilesavePath1+'/'+'Final_combined_Delay_output.xlsx'
count=0
for q in api:
    try:
        resTxt=APICall(q,'startDate','endDate',stDate,endDate,fileSavePath)
        filepath=fun_excelSave(excelFilesavePath1,xl_filecoll)
        wb = openpyxl.load_workbook(filepath)
        ws = wb['Sheet1']
        writeColumn=''
        createdRow=''
        completedRow=''
        papSpeedDays=PAPSpeedDays(ws)
        calculatedDates,increasedDate,calculatedDates1,increasedDate1=IncreasePAPDaysToCreatedDate(ws,papSpeedDays)
        aheadDelay,diffDays=CalculatedDiffCreatedDateAndEstimatedDate(ws,calculatedDates,calculatedDates1)
        UpdateIncreasedDates(ws,increasedDate)
        UpdateIncreasedDates1(ws,aheadDelay,increasedDate1)
        DiffDays(ws,diffDays)
        WriteLastColumn(ws,aheadDelay)
        wb.save(filepath)
        wb.close()
        count=count+1
        wb2=xl.load_workbook(filepath)
        ws2 = wb2.worksheets[0]
        wb3=xl.load_workbook(delay_xlfile)
        ws3=wb3.active
        mr = ws2.max_row
        mc = ws2.max_column
        AppendDelayValues(ws2,ws3,mr,mc)
        wb3.save(str(delay_xlfile))
        wb1 = openpyxl.load_workbook(excelfile)
        ws1 = wb1['Main']
        AddSheetinNewExcel(ws3,ws1,'B0','F0')
        AddSheetinNewExcel(ws3,ws1,'C0','G0')
        AddSheetinNewExcel(ws3,ws1,'E0','J0')
        AddSheetinNewExcel(ws3,ws1,'F0','K0')
        AddSheetinNewExcel(ws3,ws1,'G0','H0')
        AddSheetinNewExcel(ws3,ws1,'H0','I0')
        AddSheetinNewExcel(ws3,ws1,'I0','L0')
        AddSheetinNewExcel(ws3,ws1,'J0','M0')
        AddSheetinNewExcel(ws3,ws1,'K0','N0')
        AddSheetinNewExcel(ws3,ws1,'L0','O0')
        AddSheetinNewExcel(ws3,ws1,'M0','P0')
        AddSheetinNewExcel(ws3,ws1,'N0','Q0')
        AddSheetinNewExcel(ws3,ws1,'O0','R0')
        AddSheetinNewExcel(ws3,ws1,'P0','S0')
        AddSheetinNewExcel(ws3,ws1,'Q0','T0')
        ChangeNewXlsheetHeading()
        ws1.delete_cols(21)
        wb1.save(excelSavepath)
        wb1.close()
        wb3.close()
        if count==1:
            break
    except Exception as ex:
        print (ex)
